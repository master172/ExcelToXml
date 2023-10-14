#importing the required packages
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from xml.etree import ElementTree
from xml.dom import minidom
import datetime
#openpyxl for excel fileformat and xml.etree.ElementTree for xml

#path th xml document currently static need to make it user choosable

rowInfo = []

wb = None

ws = None

def set_indexes(path):
    global rowInfo
    global wb
    global ws
    
    wb = load_workbook(path)
    ws = wb.active
    #creating indexs for the data in the row info variable
    #looping over the first row and appending its values to cell_index to create a sort of lookup table for xml data
    #currently only iterating upto 3 colomns max for just row 1
    for row in ws.iter_rows(min_row=1, max_col=4, max_row=1,values_only=True):
        for cell in row:
            rowInfo.append(cell)

#creating values for looping over the remaining columns and rows
cell_index = 0
row_index = 0

to_write = ""
filename = "output.xml"


to_write += '<?xml version = "1.0" encoding="UTF-8"?> \n'
"""print('<?xml version = "1.0" encoding="UTF-8"?>')"""


#a function to iterate a certain row (needs revision)
def convert_row(i):
    global to_write
    #at each row we update the cell_index variable to 0
    #the cell_index stores what cell is currently selected in the current row 
    #for example if current row is row 2 and cell_index is 1 then the selected cell is B2 (second row and second column)
    global cell_index
    cell_index = 0

    #row_index stores the current iterating row used to know what entry value we are at 
    global row_index
    row_index += 1

    """print("<Entry" + str(row_index-1)+">")"""

    to_write += "   <Entry" + str(row_index-1)+">\n"
    #for each cell in the current row we print the value along with the row_index 
    #using the rowInfo to know what entry point we are at
    for j in i:
        cell_index += 1
        """print("<"+rowInfo[cell_index-1]+">",j,"</"+rowInfo[cell_index-1]+">")"""
        to_write += "          <" + rowInfo[cell_index-1] +">" + convertTuple(j) + "</" +rowInfo[cell_index-1]+">" +  "\n"

    """print("</Entry" + str(row_index-1)+">")"""
    to_write += "   </Entry" + str(row_index-1)+">\n"

def convertTuple(tup):

    if type(tup) is datetime.date or type(tup) is datetime.datetime:

        string_representation = tup.strftime("%Y-%m-%d")
        return string_representation
    elif type(tup) is int:
        return str(tup)
    elif type(tup) is float:
        return str(tup)
    string = ''
    for i in tup:
        string += i
    return string

#using a function to iterate over all rows
def convert():
    global to_write
    """print("<data>")"""
    to_write += "<data>\n"
    for i in ws.iter_rows(min_row = 2,max_col=4,max_row=5,values_only = True):
        convert_row(i)
    """print("</data>")"""
    to_write += "</data>"

def store_string_as_file(string, filename,pathToFolder):
  #Stores a string as a file.

  #Args:
  #  string: The string to store.
  #  filename: The name of the file to store the string in.
    print(pathToFolder)
    with open(pathToFolder + "\\" + filename, "w") as f:
        f.write(string)

def work(pathToFolder):
    print("\n")
    print(to_write)
    print("\n")
    store_string_as_file(to_write,filename,pathToFolder)

def start(pathToFile,pathToFolder):
    set_indexes(pathToFile)
    convert()
    work(pathToFolder)

